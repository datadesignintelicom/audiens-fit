# ---------------------------------------------------------------------------
# Audiens Fit — exportador.py
# Gera o XLSX de exportação em uma única planilha com duas abas:
# "comentarios" (universo classificado, um por linha) e "resumo"
# (indicadores, distribuição, cobertura, percepções, temas, posicionamentos).
# Arquivo único dispensa a permissão de múltiplos downloads do navegador.
#
# Criado por Daniel Bastos · Data Design Inteligência de Comunicação — MIT
# ---------------------------------------------------------------------------

import io
from datetime import datetime


def gerar_xlsx(d):
    """Recebe o dicionário de resultado da análise e devolve os bytes do XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    dt = datetime.now().strftime("%d/%m/%Y %H:%M")
    fonte = d.get("fonte", "")
    negrito = Font(bold=True)

    def cabecalho(ws, nomes):
        ws.append(nomes)
        for celula in ws[1]:
            celula.font = negrito

    # ── aba 1: universo classificado ──
    ws = wb.active
    ws.title = "comentarios"
    cabecalho(ws, ["data_hora", "fonte", "comentario", "sentimento_contextual",
                   "ironia", "via", "percepcoes_atribuidas", "tema_atribuido",
                   "posicionamento_atribuido"])
    for it in d.get("detalhes", []):
        ws.append([dt, fonte, it.get("texto", ""), it.get("sentimento", ""),
                   "sim" if it.get("ironia") else "nao", it.get("via", ""),
                   " | ".join(it.get("percepcoes", [])), it.get("tema", ""),
                   it.get("posicionamento", "")])
    ws.column_dimensions["C"].width = 80

    # ── aba 2: resumo com funil de cobertura ──
    ws = wb.create_sheet("resumo")
    cabecalho(ws, ["data_hora", "fonte", "grupo_de_dados", "rotulo",
                   "quantidade_comentarios", "base_do_percentual", "percentual"])

    def linha(grupo, rotulo, n="", base="", pct=""):
        ws.append([dt, fonte, grupo, rotulo, n, base, pct])

    m = d.get("metodologia", {})
    total = d.get("total") or 1
    linha("indicador", "receptividade_0a100", d.get("receptividade", ""))
    linha("indicador", f"perfil: {m.get('perfil', '')} / {m.get('modelo', '')}")
    for k, n in (d.get("distribuicao") or {}).items():
        linha("sentimento", k, n, total, round(n / total * 100, 1))
    for k, v in (d.get("cobertura") or {}).items():
        linha("cobertura", f"{k}_classificados", v.get("classificados", 0),
              v.get("total", 0), round(v.get("classificados", 0) / (v.get("total") or 1) * 100, 1))
    for p in d.get("percepcoes", []):
        linha("percepcao", p.get("label", ""), p.get("n", ""), total, p.get("pct", ""))
    for t in d.get("temas", []):
        linha("tema", t.get("label", ""), t.get("n", ""), "", t.get("pct", ""))
    for p in d.get("posicionamentos", []):
        linha("posicionamento", p.get("label", ""), p.get("n", ""), total, p.get("pct", ""))
    ws.column_dimensions["D"].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
