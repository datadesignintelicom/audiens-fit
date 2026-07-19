# ---------------------------------------------------------------------------
# Audiens Fit — importador.py
# Leitura de planilhas (XLSX/CSV) com detecção automática da coluna de
# comentários: por nome do cabeçalho ou, na falta, pela coluna de texto
# com maior comprimento médio.
#
# Criado por Daniel Bastos · Data Design Inteligência de Comunicação — MIT
# ---------------------------------------------------------------------------

import csv
import io
import re

_NOMES_TEXTO = re.compile(r"coment|texto|text|content|mensag|message|body|caption", re.I)


def _linhas_de_xlsx(conteudo):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    for linha in ws.iter_rows(values_only=True):
        yield ["" if c is None else str(c) for c in linha]


def _linhas_de_csv(conteudo):
    texto = conteudo.decode("utf-8-sig", errors="replace")
    amostra = texto[:4096]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=";,\t")
    except Exception:
        dialeto = csv.excel
    for linha in csv.reader(io.StringIO(texto), dialeto):
        yield linha


def extrair_comentarios(conteudo, nome_arquivo):
    """
    Retorna (lista_de_comentarios, aviso|None) ou levanta ValueError.
    """
    if nome_arquivo.lower().endswith((".xlsx", ".xlsm")):
        linhas = list(_linhas_de_xlsx(conteudo))
    elif nome_arquivo.lower().endswith((".csv", ".tsv", ".txt")):
        linhas = list(_linhas_de_csv(conteudo))
    else:
        raise ValueError("Formato não suportado. Envie .xlsx ou .csv.")

    linhas = [l for l in linhas if any(str(c).strip() for c in l)]
    if not linhas:
        raise ValueError("Planilha vazia.")

    cabecalho = [str(c).strip() for c in linhas[0]]
    corpo = linhas[1:] if len(linhas) > 1 else []

    # 1º critério: cabeçalho com nome típico de coluna de comentário
    col = next((i for i, nome in enumerate(cabecalho) if _NOMES_TEXTO.search(nome)), None)
    aviso = None

    # 2º critério: coluna com maior comprimento médio de texto
    if col is None:
        n_cols = max(len(l) for l in linhas)
        medias = []
        base = corpo if corpo else linhas
        for i in range(n_cols):
            valores = [str(l[i]) for l in base if i < len(l) and str(l[i]).strip()]
            medias.append(sum(len(v) for v in valores) / len(valores) if valores else 0)
        col = max(range(n_cols), key=lambda i: medias[i])
        aviso = f"Coluna de comentários detectada automaticamente (coluna {col + 1})."
        corpo = base  # sem cabeçalho reconhecido, considera tudo dado

    comentarios = [str(l[col]).strip() for l in corpo
                   if col < len(l) and str(l[col]).strip() and str(l[col]).strip().lower() != "nan"]
    if not comentarios:
        raise ValueError("Nenhum comentário encontrado na coluna detectada.")
    return comentarios, aviso
